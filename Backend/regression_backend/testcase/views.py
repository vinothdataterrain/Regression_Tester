from rest_framework import viewsets,status
from rest_framework.views import APIView
from rest_framework.response import Response
from asgiref.sync import sync_to_async
from playwright.async_api import async_playwright
from urllib.parse import urljoin
from rest_framework.response import Response
from django.http import HttpResponse
from .models import Project, TestCase, TestStep
from .serializers import ProjectSerializer, TestCaseSerializer
from threading import Thread
import asyncio
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from django.conf import settings
from django.db.models import Count, Avg
from datetime import datetime
from io import BytesIO
import re
import asyncio
import openpyxl
from uuid import uuid4
from playwright.sync_api import sync_playwright
import traceback
import json
import time
import contextlib
import threading
import io
import base64

class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes=[AllowAny]
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            "count": queryset.count(),
            "results": serializer.data
        })
    
class SummaryView(APIView):
    def get(self, request):
        project_count = Project.objects.count()
        testcase_count = TestCase.objects.count()
        teststeps_count = TestStep.objects.count()

        avg_steps = TestCase.objects.annotate(step_count=Count('steps')).aggregate(avg_steps=Avg('step_count'))['avg_steps'] or 0

        data = {
            "totalProjects" : project_count,
            "totalTestCases" : testcase_count,
            "totalTestSteps" : teststeps_count,
            "avgSteps" : round(avg_steps, 1)
        }
        return Response(data)

class TestCaseViewSet(viewsets.ModelViewSet):
    queryset = TestCase.objects.all()
    serializer_class = TestCaseSerializer
    permission_classes = [AllowAny]

    # --- Utility: Resolve placeholders in step values ---
    def resolve_step_value(self, step_value, row_data):
        """
        Replace placeholders like {{column}} with row_data[column] if available.
        """
        if not step_value:
            return None

        if row_data:
            matches = re.findall(r"\{\{(.*?)\}\}", step_value)
            for m in matches:
                if m in row_data and row_data[m] is not None:
                    step_value = step_value.replace(f"{{{{{m}}}}}", str(row_data[m]))
        return step_value

    # --- Core Playwright Runner ---
    async def run_steps(self, steps, project_url, row_data=None):
        results = []
        
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=False, slow_mo=200)
            page = await browser.new_page()

            for step in steps:
                try:
                    final_url = step.url or project_url
                    if step.action == "goto" and final_url and not final_url.startswith("http"):
                        final_url = urljoin(project_url, step.url)

                    # Resolve value (JSON value or Excel/CSV override)
                    step_value = self.resolve_step_value(step.value, row_data)

                    # --- Action Handlers ---
                    if step.action == "goto":
                        await page.goto(final_url)

                    elif step.action == "fill":
                        if not step.selector:
                            raise ValueError("Selector required for fill")
                        await page.fill(step.selector, step_value or "")

                    elif step.action == "click":
                        if not step.selector:
                            raise ValueError("Selector required for click")
                        await page.click(step.selector)

                    elif step.action == "select":
                        if not step.selector or not step_value:
                            raise ValueError("Selector & value required for select")
                        await page.select_option(step.selector, step_value)

                    elif step.action == "check":
                        if not step.selector:
                            raise ValueError("Selector required for check")
                        await page.check(step.selector)

                    elif step.action == "uncheck":
                        if not step.selector:
                            raise ValueError("Selector required for uncheck")
                        await page.uncheck(step.selector)

                    elif step.action == "assert":
                        if not step.selector:
                            raise ValueError("Selector required for assert")
                        element_text = await page.text_content(step.selector) or ""
                        expected_value = step_value or ""
                        if expected_value not in element_text:
                            raise AssertionError(
                                f"Expected '{expected_value}' in '{element_text}'"
                            )

                    elif step.action == "expect_visible":
                        if not step.selector:
                            raise ValueError("Selector required for expect_visible")
                        await page.wait_for_selector(step.selector, state="visible", timeout=5000)

                    elif step.action == "expect_hidden":
                        if not step.selector:
                            raise ValueError("Selector required for expect_hidden")
                        await page.wait_for_selector(step.selector, state="hidden", timeout=5000)

                    elif step.action == "expect_url":
                        expected_value = step_value or ""
                        if expected_value not in page.url:
                            raise AssertionError(
                                f"Expected URL to contain '{expected_value}', got '{page.url}'"
                            )

                    elif step.action == "expect_title":
                        expected_value = step_value or ""
                        title = await page.title()
                        if expected_value not in title:
                            raise AssertionError(
                                f"Expected title '{expected_value}' in '{title}'"
                            )   
                    elif step.action == "wait":
                            if step.selector:
                                timeout = int(step.value) if step.value else 5000
                                await page.wait_for_selector(step.selector, timeout=timeout)
                            elif step.value:
                                timeout = int(step.value)
                                await page.wait_for_timeout(timeout = timeout)
                            else:
                                raise ValueError("Wait action requires either selector or value")
                    else:
                        raise ValueError(f"Unknown action {step.action}")

                    results.append({
                        "action": step.action,
                        "value": step.value,
                        "step_number": step.order + 1,
                        "status": "passed"
                    })

                except Exception as e:
                    results.append({
                        "action": step.action,
                        "value": step.value,
                        "step_number": step.order + 1,
                        "status": "failed",
                        "error": str(e),
                    })
                    break  # stop on first failure

            await browser.close()
        return results

    # --- API Endpoint ---
    @action(detail=True, methods=["post"])
    def run(self, request, pk=None):
        testcase = self.get_object()
        steps = list(testcase.steps.all().order_by("order", "id"))
        print(steps,"Stepss")
        project_url = testcase.project.url if testcase.project else ""

        # --- Case 1: File uploaded (Excel or CSV) ---
        if "file" in request.FILES:
            file = request.FILES["file"]
            # Load uploaded Excel
            wb_input = openpyxl.load_workbook(file)
            ws_input = wb_input.active
            headers = [cell.value for cell in ws_input[1]]

            # Prepare output workbook in memory
            wb_output = openpyxl.Workbook()
            ws_output = wb_output.active
            ws_output.append(headers + ["Status", "Step Results"])  # header row

            for row in ws_input.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                row_results = asyncio.run(self.run_steps(steps, project_url, row_data))
                overall_status = "passed" if all(r["status"] == "passed" for r in row_results) else "failed"

                # Save row + results to output sheet
                results_str = "; ".join([f"{r['step_number']}:{r['action']}={r['status']}" for r in row_results])
                ws_output.append(list(row) + [overall_status, results_str])

            # Save to in-memory bytes
            output = BytesIO()
            wb_output.save(output)
            output.seek(0)

            # Return as downloadable Excel
            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename=testcase_{testcase.name}_results.xlsx'
            return response

        # --- Case 2: Default run (use JSON values directly) ---
        else:
            results = asyncio.run(self.run_steps(steps, project_url))
            overall_status = "passed" if all(r["status"] == "passed" for r in results) else "failed"

            return Response({
                "testcase": testcase.id,
                "mode": "default",
                "status": overall_status,
                "results": results
            }, status=status.HTTP_200_OK)
        
# class PlaywrightRunView(APIView):
    # def post(self, request, *args, **kwargs):
    #     script_content = request.data.get('script')
    #     if not script_content:
    #         return Response({"error" : "Script is required"}, status=status.HTTP_400_BAD_REQUEST)
    #     tmp_file_path = None

    #     try:
    #         with tempfile.NamedTemporaryFile(delete= False, suffix=".py") as tmp_file:
    #             tmp_file.write(script_content.encode())
    #             tmp_file_path = tmp_file.name

    #             python_executable = sys.executable

    #             process = subprocess.Popen(
    #                 [python_executable,tmp_file_path],
    #                 stdout = subprocess.PIPE,
    #                 stderr = subprocess.PIPE,
    #                 text = True,
    #                 cwd =settings.BASE_DIR
    #             )

    #             try:
    #                 stdout, stderr = process.communicate(timeout=30)
    #                 return_code = process.returncode
    #             except subprocess.TimeoutExpired:
    #                 process.kill()
    #                 return Response({"error" : "subprocess time exceeded"}, status=status.HTTP_408_REQUEST_TIMEOUT)
                
    #             test_status = "passed" if return_code == 0  else "failed"

    #             response_data = {
    #                 "status" : test_status,
    #                 "stdout" : stdout.strip(),
    #                 "stderr" : stderr.strip(),
    #                 "return_code" : return_code
    #             }

    #             return Response(response_data, status=status.HTTP_200_OK)
    #     except Exception as e:
    #         return Response({"status" : "failed", "error" : f"An error occured : {str(e)}"}, status = status.HTTP_500_INTERNAL_SERVER_ERROR)
    #     finally:
    #         if tmp_file_path and os.path.exists(tmp_file_path):
    #             os.remove(tmp_file_path)    
         
class PlaywrightExecutorWithScreenshots:
    def __init__(self):
        self.logs = []
        self.screenshots = []
        self.start_time = None
        self.end_time = None
        self.current_page = None
    
    def log(self, message):
        """Enhanced logging function"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        self.logs.append(log_entry)
    
    def take_screenshot(self, page=None, description="Screenshot"):
        """Capture screenshot and convert to base64 for frontend display"""
        try:
            # Use provided page or current page
            screenshot_page = page or self.current_page
            if not screenshot_page:
                self.log("⚠️ No page available for screenshot")
                return
            
            # Take screenshot as bytes
            screenshot_bytes = screenshot_page.screenshot(full_page=True)
            
            # Convert to base64 for JSON response
            screenshot_base64 = base64.b64encode(screenshot_bytes).decode('utf-8')
            
            screenshot_info = {
                'description': description,
                'timestamp': datetime.now().isoformat(),
                'data': screenshot_base64,
                'url': screenshot_page.url if screenshot_page else 'unknown'
            }
            
            self.screenshots.append(screenshot_info)
            self.log(f"📸 Screenshot captured: {description}")
            
        except Exception as e:
            self.log(f"❌ Failed to capture screenshot: {str(e)}")
    
    def auto_screenshot_wrapper(self, func, page, description):
        """Wrapper to automatically take screenshots after actions"""
        try:
            result = func()
            # Wait a moment for any animations/loading
            page.wait_for_timeout(500)
            self.take_screenshot(page, description)
            return result
        except Exception as e:
            self.take_screenshot(page, f"Error during: {description}")
            raise e
    
    def execute_script(self, script_content, timeout=60, auto_screenshots=True):
        """Execute Playwright script with screenshot capture"""
        self.start_time = datetime.now()
        self.logs = []
        self.screenshots = []
        
        # Enhanced execution environment with screenshot functions
        def enhanced_take_screenshot(page=None, description="Manual Screenshot"):
            self.take_screenshot(page or self.current_page, description)
        
        def enhanced_goto(page, url, description=None):
            self.log(f"🌐 Navigating to: {url}")
            page.goto(url)
            page.wait_for_load_state('networkidle')
            if auto_screenshots:
                desc = description or f"Navigation to {url}"
                self.take_screenshot(page, desc)
        
        def enhanced_fill(page, selector, value, description=None):
            self.log(f"✏️ Filling '{selector}' with value")
            page.fill(selector, value)
            if auto_screenshots:
                desc = description or f"Filled {selector}"
                self.take_screenshot(page, desc)
        
        def enhanced_click(page, selector, description=None):
            self.log(f"👆 Clicking: {selector}")
            page.click(selector)
            page.wait_for_timeout(500)  # Wait for any animations
            if auto_screenshots:
                desc = description or f"Clicked {selector}"
                self.take_screenshot(page, desc)
        
        script_globals = {
            '__builtins__': __builtins__,
            'sync_playwright': sync_playwright,
            'log': self.log,
            'print': self.log,
            'take_screenshot': enhanced_take_screenshot,
            'enhanced_goto': enhanced_goto,
            'enhanced_fill': enhanced_fill,
            'enhanced_click': enhanced_click,
            'datetime': datetime,
            'time': time,
            'json': json,
        }
        
        # Capture stdout/stderr
        stdout_capture = io.StringIO()
        stderr_capture = io.StringIO()
        
        try:
            self.log("🚀 Starting Playwright script execution with screenshots...")
            
            execution_error = None
            
            def run_script():
                nonlocal execution_error
                try:
                    with contextlib.redirect_stdout(stdout_capture), \
                         contextlib.redirect_stderr(stderr_capture):
                        exec(script_content, script_globals)
                except Exception as e:
                    execution_error = e
            
            # Execute with timeout
            thread = threading.Thread(target=run_script)
            thread.daemon = True
            thread.start()
            thread.join(timeout=timeout)
            
            if thread.is_alive():
                raise TimeoutError(f"Script execution timeout after {timeout} seconds")
            
            if execution_error:
                raise execution_error
            
            self.end_time = datetime.now()
            self.log("✅ Script execution completed successfully!")
            
            return {
                'status': 'passed',
                'message': 'Script executed successfully with screenshots captured',
                'logs': self.logs,
                'screenshots': self.screenshots,
                'screenshot_count': len(self.screenshots),
                'stdout': stdout_capture.getvalue(),
                'stderr': stderr_capture.getvalue(),
                'execution_time': (self.end_time - self.start_time).total_seconds(),
                'start_time': self.start_time.isoformat(),
                'end_time': self.end_time.isoformat()
            }
            
        except Exception as e:
            self.end_time = datetime.now()
            error_msg = str(e)
            traceback_str = traceback.format_exc()
            
            self.log(f"❌ Script execution failed: {error_msg}")
            
            return {
                'status': 'failed',
                'message': f'Script execution failed: {error_msg}',
                'error': error_msg,
                'traceback': traceback_str,
                'logs': self.logs,
                'screenshots': self.screenshots,
                'screenshot_count': len(self.screenshots),
                'stdout': stdout_capture.getvalue(),
                'stderr': stderr_capture.getvalue(),
                'execution_time': (self.end_time - self.start_time).total_seconds() if self.start_time else 0,
                'start_time': self.start_time.isoformat() if self.start_time else None,
                'end_time': self.end_time.isoformat() if self.end_time else None
            }

class PlaywrightRunView(APIView):
    def post(self, request, *args, **kwargs):
        script_content = request.data.get('script')
        timeout = int(request.data.get('timeout', 60))
        auto_screenshots = request.data.get('auto_screenshots', True)
        
        if not script_content:
            return Response({
                "error": "Script content is required"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            executor = PlaywrightExecutorWithScreenshots()
            result = executor.execute_script(script_content, timeout, auto_screenshots)
            
            return Response(result, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                "status": "failed",
                "error": f"Unexpected error: {str(e)}",
                "traceback": traceback.format_exc()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# class TestCaseViewSet(viewsets.ModelViewSet):
#     queryset = TestCase.objects.all()
#     serializer_class = TestCaseSerializer
#     permission_classes = [AllowAny]

#     # def create(self, request, *args, **kwargs):
#     #     try:
#     #         print("glkdfjd")
#     #         return Response({'msg':"fldk"}, status=status.HTTP_200_OK)
#     #     except Exception as e:
#     #         return Response({"error", str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#     @action(detail=True, methods=["post"])
#     def run(self, request, pk=None):
#         testcase = self.get_object()
#         steps = list(testcase.steps.all().order_by("order", "id"))
#         project_url = testcase.project.url

#         async def run_playwright():
#             results = []
#             async with async_playwright() as p:
#                 browser = await p.chromium.launch(headless=False, slow_mo=500)
#                 page = await browser.new_page()
#                 for step in steps:
#                     try:
#                         final_url = step.url or project_url
#                         if step.action == "goto" and not final_url.startswith("http"):
#                             from urllib.parse import urljoin
#                             final_url = urljoin(project_url, step.url)

#                         # Navigation
#                         if step.action == "goto":
#                             await page.goto(final_url)

#                         # Fill inputs
#                         elif step.action == "fill":
#                             if not step.selector:
#                                 raise ValueError("Selector required for fill")
#                             await page.fill(step.selector, step.value or "")

#                         # Click
#                         elif step.action == "click":
#                             if not step.selector:
#                                 raise ValueError("Selector required for click")
#                             await page.click(step.selector)

#                         # Select dropdown option
#                         elif step.action == "select":
#                             if not step.selector:
#                                 raise ValueError("Selector required for select")
#                             if not step.value:
#                                 raise ValueError("Value required for select")
#                             await page.select_option(step.selector, step.value)

#                         # Check a checkbox
#                         elif step.action == "check":
#                             if not step.selector:
#                                 raise ValueError("Selector required for check")
#                             await page.check(step.selector)

#                         # Uncheck a checkbox
#                         elif step.action == "uncheck":
#                             if not step.selector:
#                                 raise ValueError("Selector required for uncheck")
#                             await page.uncheck(step.selector)

#                         # Assert element text contains
#                         elif step.action == "assert":
#                             if not step.selector:
#                                 raise ValueError("Selector required for assert")
#                             element_text = await page.text_content(step.selector) or ""
#                             expected_value = step.value or ""
#                             if expected_value not in element_text:
#                                 raise AssertionError(
#                                     f"Expected '{expected_value}' in '{element_text}'"
#                                 )

#                         # Expect element to be visible
#                         elif step.action == "expect_visible":
#                             if not step.selector:
#                                 raise ValueError("Selector required for expect_visible")
#                             await page.wait_for_selector(step.selector, state="visible", timeout=5000)

#                         # Expect element to be hidden
#                         elif step.action == "expect_hidden":
#                             if not step.selector:
#                                 raise ValueError("Selector required for expect_hidden")
#                             await page.wait_for_selector(step.selector, state="hidden", timeout=5000)

#                         # Expect URL contains
#                         elif step.action == "expect_url":
#                             expected_value = step.value or ""
#                             current_url = page.url
#                             if expected_value not in current_url:
#                                 raise AssertionError(
#                                     f"Expected URL to contain '{expected_value}', got '{current_url}'"
#                                 )

#                         # Expect title contains
#                         elif step.action == "expect_title":
#                             expected_value = step.value or ""
#                             title = await page.title()
#                             if expected_value not in title:
#                                 raise AssertionError(
#                                     f"Expected title '{expected_value}' in '{title}'"
#                                 )

#                         # wait for an action
#                         elif step.action == "wait":
#                             if step.selector:
#                                 timeout = int(step.value) if step.value else 5000
#                                 await page.wait_for_selector(step.selector, timeout=timeout)
#                             elif step.value:
#                                 timeout = int(step.value)
#                                 await page.wait_for_timeout(timeout = timeout)
#                             else:
#                                 raise ValueError("Wait action requires either selector or value")

#                         else:
#                             raise ValueError(f"Unknown action {step.action}")

#                         # ✅ Success
#                         results.append({"action": step.action,
#                             "value": step.value,"step_number": step.order + 1, "status": "passed"})

#                     except Exception as e:
#                         results.append({
#                             "action": step.action,
#                             "value": step.value,
#                             "step_number": step.order + 1,
#                             "status": "failed",
#                             "error": str(e),
#                         })
#                         break

#                 await browser.close()
#             return results

#         results = asyncio.run(run_playwright())
#         overall_status = "passed"
#         for step in results:
#             if step["status"] == "failed":
#                 overall_status = "failed"
#                 break 
#         return Response({"testcase": testcase.id, "status": overall_status , "results": results})