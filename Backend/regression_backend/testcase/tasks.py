import os
import tempfile
import openpyxl
from concurrent.futures import ThreadPoolExecutor
from django.conf import settings
from .models import TestRun, TestCase
import asyncio
from .playwright_runner import run_testcase_async  # your function

executor = ThreadPoolExecutor(max_workers=3)

def process_testcase(test_run_id, input_file_path=None):
    """
    Background thread function to run Playwright tests and update TestRun.
    """
    test_run = TestRun.objects.get(id=test_run_id)
    testcase = test_run.testcase

    test_run.status = "running"
    test_run.progress = "0%"
    test_run.save()

    try:
        if input_file_path:
            wb = openpyxl.load_workbook(input_file_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = []

        # Add result column if not exists
        result_col = len(headers) + 1
        ws.cell(row=1, column=result_col, value="Result")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))
            try:
                steps = list(testcase.steps.all().values()) # JSONField from model
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                results = loop.run_until_complete(run_testcase_async(steps, values=row_data))
                step_results = [r for r in results if "action" in r and "status" in r]
                print(results)
                pretty_results = "\n".join(
                    [f"{r['action']} | {r['value']} | {r['status']}, " for r in step_results]
                    )
                row_status = "PASS" if all(r["status"] == "passed" for r in step_results) else "FAIL"
                print(row_status)
                print(pretty_results)
                ws.cell(row=row_idx, column=result_col, value=row_status)
                ws.cell(row=row_idx, column=result_col + 1, value=pretty_results)
                video_info = results[-1]["video"] if "video" in results[-1] else None
                if video_info:
                    ws.cell(row=row_idx, column=result_col + 2, value=video_info)

            except Exception as e:
                ws.cell(row=row_idx, column=result_col, value=f"FAIL: {str(e)}")

        # Save results file
        results_dir = os.path.join(settings.MEDIA_ROOT, "results")
        os.makedirs(results_dir, exist_ok=True)
        result_path = os.path.join(results_dir, f"testcase_{testcase.id}_latest.xlsx")
        wb.save(result_path)

        test_run.status = "completed"
        test_run.result_file.name = f"results/testcase_{testcase.id}_latest.xlsx"
        test_run.progress = "100%"
        test_run.save()

    except Exception as e:
        test_run.status = "failed"
        test_run.progress = str(e)
        test_run.save()


def run_testcase_in_background(testcase, input_file=None):
    """
    Entry point to run TestCase in background thread.
    """
    test_run = TestRun.objects.create(testcase=testcase, status="queued", progress="0%")

    input_file_path = None
    if input_file:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in input_file.chunks():
                tmp.write(chunk)
            input_file_path = tmp.name

    executor.submit(process_testcase, test_run.id, input_file_path)
    return test_run